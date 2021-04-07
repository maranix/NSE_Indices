import openpyxl
wb = openpyxl.Workbook()
Sheet_name = wb.sheetnames
wb.save(filename='IndiciesHistoricaldata.xlsx')
wb1 = openpyxl.load_workbook('IndiciesHistoricaldata.xlsx')
wb2 = openpyxl.load_workbook('NIFTY%20AUTO.xlsx')
wb3 = openpyxl.load_workbook('NIFTY%20BANK.xlsx')
wb4 = openpyxl.load_workbook('NIFTY%20CONSUMER%20DURABLES.xlsx')
wb5 = openpyxl.load_workbook('NIFTY%20FIN%20SERVICE.xlsx')
wb6 = openpyxl.load_workbook('NIFTY%20FINSRV25%2050.xlsx')
wb7 = openpyxl.load_workbook('NIFTY%20FMCG.xlsx')
wb8 = openpyxl.load_workbook('NIFTY%20Healthcare%20Index.xlsx')
wb9 = openpyxl.load_workbook('NIFTY%20IT.xlsx')
wb10 = openpyxl.load_workbook('NIFTY%20MEDIA.xlsx')
wb11 = openpyxl.load_workbook('NIFTY%20METAL.xlsx')
wb12 = openpyxl.load_workbook('NIFTY%20OIL%20%26%20GAS.xlsx')
wb13 = openpyxl.load_workbook('NIFTY%20PHARMA.xlsx')
wb14 = openpyxl.load_workbook('NIFTY%20PSU%20BANK.xlsx')
wb15 = openpyxl.load_workbook('NIFTY%20PVT%20BANK.xlsx')
wb16 = openpyxl.load_workbook('NIFTY%20REALTY.xlsx')
sheet2 = wb1['Sheet']
sheet1 = wb2['Sheet1']
def new_func(wb3, wb4, wb5, wb6, wb7, wb8, wb9, wb10, wb11, wb12, wb13, wb14, wb15, wb16):
    sheet3 = wb3['Sheet1']
    sheet4 = wb4['Sheet1']
    sheet5 = wb5['Sheet1']
    sheet6 = wb6['Sheet1']
    sheet7 = wb7['Sheet1']
    sheet8 = wb8['Sheet1']
    sheet9 = wb9['Sheet1']
    sheet10 = wb10['Sheet1']
    sheet11 = wb11['Sheet1']
    sheet12 = wb12['Sheet1']
    sheet13 = wb13['Sheet1']
    sheet14 = wb14['Sheet1']
    sheet15 = wb15['Sheet1']
    sheet16 = wb16['Sheet1']
    return sheet3,sheet4,sheet5,sheet6,sheet7,sheet8,sheet9,sheet10,sheet11,sheet12,sheet13,sheet14,sheet15,sheet16

sheet3, sheet4, sheet5, sheet6, sheet7, sheet8, sheet9, sheet10, sheet11, sheet12, sheet13, sheet14, sheet15, sheet16 = new_func(wb3, wb4, wb5, wb6, wb7, wb8, wb9, wb10, wb11, wb12, wb13, wb14, wb15, wb16)

for i,row in enumerate(sheet1.iter_rows()):
    for j,col in enumerate(row):
        sheet2.cell(row=i+1,column=j+1).value = col.value
for i,row in enumerate(sheet3.iter_rows()):
    for j,col in enumerate(row):
        sheet2.cell(row=i+1,column=j+12).value = col.value
for i,row in enumerate(sheet4.iter_rows()):
    for j,col in enumerate(row):
        sheet2.cell(row=i+1,column=j+23).value = col.value
for i,row in enumerate(sheet5.iter_rows()):
    for j,col in enumerate(row):
        sheet2.cell(row=i+1,column=j+34).value = col.value
for i,row in enumerate(sheet6.iter_rows()):
    for j,col in enumerate(row):
        sheet2.cell(row=i+1,column=j+45).value = col.value
for i,row in enumerate(sheet7.iter_rows()):
    for j,col in enumerate(row):
        sheet2.cell(row=i+1,column=j+56).value = col.value
for i,row in enumerate(sheet8.iter_rows()):
    for j,col in enumerate(row):
        sheet2.cell(row=i+1,column=j+67).value = col.value
for i,row in enumerate(sheet9.iter_rows()):
    for j,col in enumerate(row):
        sheet2.cell(row=i+1,column=j+78).value = col.value
for i,row in enumerate(sheet10.iter_rows()):
    for j,col in enumerate(row):
        sheet2.cell(row=i+1,column=j+89).value = col.value
for i,row in enumerate(sheet11.iter_rows()):
    for j,col in enumerate(row):
        sheet2.cell(row=i+1,column=j+100).value = col.value
for i,row in enumerate(sheet12.iter_rows()):
    for j,col in enumerate(row):
        sheet2.cell(row=i+1,column=j+111).value = col.value
for i,row in enumerate(sheet13.iter_rows()):
    for j,col in enumerate(row):
        sheet2.cell(row=i+1,column=j+122).value = col.value
for i,row in enumerate(sheet14.iter_rows()):
    for j,col in enumerate(row):
        sheet2.cell(row=i+1,column=j+133).value = col.value
for i,row in enumerate(sheet15.iter_rows()):
    for j,col in enumerate(row):
        sheet2.cell(row=i+1,column=j+144).value = col.value
for i,row in enumerate(sheet16.iter_rows()):
    for j,col in enumerate(row):
        sheet2.cell(row=i+1,column=j+155).value = col.value
wb1.save('IndiciesHistoricaldata.xlsx')